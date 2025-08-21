from openai import OpenAI
import pandas as pd
import gspread
from gspread_dataframe import set_with_dataframe
from oauth2client.service_account import ServiceAccountCredentials
from collections import OrderedDict
import json
import os
import glob
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
from openpyxl import load_workbook
from dotenv import load_dotenv
import shutil
from pathlib import Path
from tqdm import tqdm
from datetime import datetime, timedelta

def write_to_gsheet(sheet, data, cell_map):
    """
    Writes values from row_data to specific cells in the sheet using cell_map.
    Used to write 'AT WIP' gsheet values into custom PO gsheet.

    Args:
        sheet (gsheet object): gspread worksheet to write into.
        data (dict): mapping of field names (e.g. "CompanyPO#") and their values.
        cell_map (dict): mapping of field names to cell addresses.
    
    Returns:
        None
    """
    updates = []
    for field, cell in cell_map.items():
        value = data.get(field, "")
        updates.append({
            'range': cell,
            'values': [[value]]
        })

    sheet.batch_update(updates)

def write_to_excel(file_pth, PO_json, cell_map):
    """
    Write data to an existing Excel file using cell mappings.
    Used to write location PO info from 'AT WIP' to previously saved custom PO excel.

    Args:
        file_path (str): Path to the Excel file.
        purchase_order_json (dict): Data to write (e.g., {"Company PO#": "12345", "0": 10}).
        cell_map (dict): mapping of field names to cell addresses (e.g., {"Company PO#": "G6", "0": "C12"}).
    """
    try:
        # Load the existing workbookt
        try:
            workbook = load_workbook(file_pth)
        except FileNotFoundError:
            raise ValueError(
                f"\nFile {file_pth} not found. The file does not exist in the given path. "
                "It may have been deleted or the MASTER pdf may have not been processed yet."
            )
        
        # Access the worksheet (there should be only one sheet (tab) named after the size)
        tab = workbook.active
        
        # Write data to the specified cells
        for key, value in PO_json.items():
            if key in cell_map:
                cell_address = cell_map[key]
                tab[cell_address] = value
        
        # Save the workbook
        workbook.save(file_pth)
        #print(f"Data successfully written to {file_pth}.")
        
    except Exception as e:
        print(f"\nError writing to Excel: {e}")
        input("\nDelete the row in 'AT WIP' google sheet and rerun script with the same pdf. Press enter to exit.")   
        exit()

def get_consistent_output(client, client_input, model, temperature, use=False):
    """Call the model twice and only accept consistent, valid output."""
    for _ in range(3):  # Try up to 3 times
        out1 = client.responses.create(model=model, temperature=temperature, input=client_input)

        if not use:
            return out1.output_text

        out2 = client.responses.create(model=model, temperature=temperature, input=client_input)
        
        if out1.output_text == out2.output_text:
            return out1.output_text
    raise ValueError("Model output inconsistency detected! Try again.")

def process_PDF(sheet, client_openai, client_input, openai_config):
    """
    Runs openai's model to process a pdf and the prompt to extract
    necessary data as a json and write into the given gsheet.
    Used to write pdf data into 'AT WIP'.

    Args
    ----
    sheet : gsheet object
        Google sheet to input pdf data to.
    client_open : openai object 
        OpenAI(api_key=OPENAI_API_KEY).
    client_input : dict
        Input to the openai model.
    openai_config : dict
        Configs like model and temperature.
    
    Note
    ----
    Output of openai model need to be formatted to match headers
    in 'AT WIP'. 
    
    Example of openai model output vs. 'AT WIP' column headers:
    
    model_output = [
        "STYLE#", "Color", "Order#", "Location", "Size {i}"'s, 
        "Total", "PO ISSUE", "Requested Ship Date", "Inseam Size",
        "DESCRIPTION", "C Price"
    ]

    necessary_headers = [
        "STYLE#", "Color", "Company PO#", "Location", "Vendor", "{i}"'s,
        "Total", "Ordered QTY", "PO ISSUE", "CXL", "Inseam", "DESCRIPTION", 
        "C Price"
    ]

    - Extract Company PO# and Location from Order#
    - Change Requested Ship Date to CXL
    - Change Inseam Size to Inseam
    - Change 'Size {i}' to '{i}'; 16 -> 16/1x, 18 -> 18/2x, 20 -> 20/3x
        - Change 'Size 00' to '00'

    Returns:
        openai model's output as json "response_json"
    """
    # Get output from openai model 
    response = get_consistent_output(client_openai, client_input, **openai_config) # Set use=True for inconsistent output
    response_json = json.loads(response)
    
    # Change response_json keys to match column names in 'AT WIP' google sheet
    response_json['Company PO#'] = response_json.pop('Order#', '').split("-")[0]
    response_json['CXL'] = response_json.pop('Requested Ship Date', '')
    response_json['Inseam'] = response_json.pop('Inseam Size', '')
    response_json['Location'] = response_json.pop('Location', '').upper()

    for size in range(0, 26, 2):
        if size == 16:
            response_json[f'{size} / 1X'] = response_json.pop(f'Size {size}', '')
        elif size == 18:
            response_json[f'{size} / 2X'] = response_json.pop(f'Size {size}', '')
        elif size == 20:
            response_json[f'{size} / 3X'] = response_json.pop(f'Size {size}', '')
        else:
            response_json[f'{size}'] = response_json.pop(f'Size {size}', '')
    response_json['00'] = response_json.pop('Size 00', '')
    
    # Retrieve column headers from 'AT WIP'
    required_cols = sheet.row_values(1)
    required_cols = [word.strip() for word in required_cols] # 
    
    # Create dataframe with all column headers and with response_json values for shared fields
    df = pd.DataFrame([response_json])
    df = df.astype(str)
    df = df.reindex(columns=required_cols)

    next_row = len(sheet.get_all_values()) + 1 # find next empty row

    set_with_dataframe(sheet, df, row=next_row, include_column_header=False)

    return response_json

def create_custom_PO(sheet, response_json, creds, tab_names):
    """
    Maps necessary 'AT WIP' gsheet values to custom PO gsheet.
    Custom PO excel is saved only for MASTER PDF. 

    Location PDFs GA1, PL1, AZ1 are written into it at later 
    dates. We assume MASTER always comes first.

    Notes
    -----
    PDFs of different sizes have slightly different formats. Map accordingly.

    MASTER and location PDFs are written into different rows. 
    
    Args
    ----
    sheet : gsheet object
        Google sheet to input 'AT WIP' data to.
    response_json : dict
        Output of openai model with all necessary values.
    creds : google creds
        Google drive API credentials to access the API for exports gsheets as excel.

    Returns
    -------
        Destination (directory path where custom PO excel is saved)
    """

    size = response_json["STYLE#"].split("-")[-1]
    location = response_json["Location"]
    full_style_num = response_json['STYLE#']
    style_num = full_style_num
    if size in ['PT', 'P']:
        style_num = response_json['STYLE#'].rsplit('-', 1)[0] # style# without size tag
    company_PO = response_json['Company PO#']

    # Directory path to store pdfs and excel
    destination = f"../{style_num}/PO/{company_PO}_{full_style_num}"
    
    # Check that Master file is created before any other location PDFs
    if not os.path.exists(destination):
        try:
            if location != "MASTER":
                raise ValueError(
                    "\nEnsure that location PDFs ('AZ1', 'PL1', 'GA1') are processed only after 'MASTER'.\n\n"
                    f"Current PO details:\n"
                    f"  - style_num: {full_style_num}\n"
                    f"  - order_num: {company_PO}-{location}\n\n"
                    "If the above PO details are incorrect, delete the row in 'AT WIP' google sheet and simply run the process again.\n"
                    "If they are correct, there are two scenarios:\n"
                    "   - 1) 'MASTER' PDF has not been processed yet. This would happen if you forgot to run the script with the 'MASTER PDF'.\n"
                    "        Solution: Put the corresponding 'MASTER' PDF into 'PO_pdfs' and run the script again.\n" 
                    "   - 2) 'MASTER' PDF may have been processed earlier with the wrong style number. This may happen because the" 
                    " AI that this script uses got the style# wrong for the 'MASTER' PDF.\n"
                    "        Solution: Follow the instruction in README.txt.\n"
                )
        except ValueError as e:
            print("\nAn error occurred:")
            print(e)
            input("For both cases, make sure to erase the row in 'AT WIP' google sheet. Press enter to exit.")   
            exit()
            
        os.makedirs(destination, exist_ok=False)
    
    if size not in ['PT', 'P']:
        size = 'reg'
    
    # Check if item size description '00' exists to set correct gsheet tab.
    if response_json['00']:
        if size == 'PT':
            size = 'PT_00'
        elif size == 'reg':
            size = 'reg_00'

    # Fields that are required from MASTER PDFs of all sizes.
    common_fields = {
        "Company PO#": "G6",
        "DESCRIPTION": "A10",
        "STYLE#": "A11",
        "PO ISSUE": "M6",
        "CXL": "M8",
        "Color": "A12",
        "Total": "M25",
        #"C Price": "N12"
    }

    # Create mapping to columns for each size.
    size_col_map = {
        'reg': {"0": "C", "2": "D", "4": "E", "6": "F", "8": "G", "10": "H", "12": "I", "14": "J", "16": "K"},
        'reg_00': {"00": "C", "0": "D", "2": "E", "4": "F", "6": "G", "8": "H", "10": "I", "12": "J", "14": "K", "16": "L"},
        'PT': {"0": "C", "2": "D", "4": "E", "6": "F", "8": "G", "10": "H", "12": "I", "14": "J"},
        'PT_00': {"00": "C", "0": "D", "2": "E", "4": "F", "6": "G", "8": "H", "10": "I", "12": "J", "14": "K"},
        'P': {"14": "C", "16": "D", "18": "E", "20": "F", "22": "G", "24": "H"}
    }

    # Create mapping to rows based on location
    loc_row_map = {
        'MASTER': 12,
        'GA1': 14,
        'AZ1': 16,
        'PL1': 18 
    }
    row = loc_row_map[location]

    # Create complete cell mappings
    """ 
    Example for GA1:
    cell_maps = {
        'reg': {
            "0": "C14", "2": "D14", "4": "E14", "6": "F14",
            "8": "G14", "10": "H14", "12": "I14", "14": "J14", "16": "K14"
        },
        'PT': {
            "0": "C14", "2": "D14", "4": "E14", "6": "F14",
            "8": "G14", "10": "H14", "14": "I14", "14": "J14"
        },
        'P': {
            "14": "C14", "16": "D14", "18": "E14", "20": "F14",
            "22": "G14", "24": "H14"
        }
    }
    """
    cell_maps = {
        size: {
            **(common_fields if location == "MASTER" else {}),
            **{key: f"{col}{row}" for key, col in col_map.items()}
        } for size, col_map in size_col_map.items()
    }
    
    PO_cols = cell_maps[size].keys()
    response_json['16'] = response_json.pop('16 / 1X', '')
    response_json['18'] = response_json.pop('18 / 2X', '')
    response_json['20'] = response_json.pop('20 / 3X', '')
    
    PO_json = {key: response_json[key] for key in PO_cols if key in response_json}

    # Write into gsheets
    sheet_PO = sheet[size]
    cell_map = cell_maps[size]
    sheet_name = tab_names[size]
    
    # Only Master pdf is saved as excel. Location pdfs write into the saved excel.
    if location == "MASTER":

        write_to_gsheet(sheet_PO, PO_json, cell_map)

        """Download as Excel .xlsx"""
        file_id = sheet_PO.spreadsheet.id
        sheet_to_keep = sheet_name  # the tab name you want to retain

        # Set output path
        output_path = f"{destination}/PO_{company_PO}-{location}.xlsx" 

        # Download full spreadsheet as Excel
        drive_service = build('drive', 'v3', credentials=creds)
        request = drive_service.files().export_media(
            fileId=file_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        fh = io.FileIO(output_path, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
            #print("Download progress: {:.0f}%".format(status.progress() * 100))
        fh.close()

        # Clear gsheet cells
        blank_json = {key: '' for key in PO_cols if key in response_json}
        write_to_gsheet(sheet_PO, blank_json, cell_map)

        # Remove extra sheets (tabs)
        wb = load_workbook(output_path)
        for s in wb.sheetnames:
            if s != sheet_to_keep:
                wb.remove(wb[s])
        wb.save(output_path)

    else:
        # check if PO excel exists from Master
        file_pth=os.path.join(destination, f"PO_{company_PO}-MASTER.xlsx")#path to .xlsx
        write_to_excel(file_pth, PO_json, cell_map)
    
    print(f"\nPurchase order with order# {company_PO}-{location} and style# {full_style_num} was processed. "
          "Make sure to double-check the style#.")

    return destination



if __name__ == '__main__':
    
    # Load from .env in current directory
    load_dotenv()
    
    # Load keys and directory paths
    OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
    MODEL = os.getenv("MODEL")
    #TEMPERATURE = os.getenv("TEMPERATURE")
    GOOGLE_JSON_KEY = os.getenv("GOOGLE_JSON_KEY")
    GSHEET_MAIN, GSHEET_MAIN_TAB = os.getenv("GSHEET_MAIN"), os.getenv("GSHEET_MAIN_TAB")
    GSHEET_PO = os.getenv("GSHEET_PO")
    #ROOT_DIR = os.getenv("ROOT_DIR")
    PO_DIR = os.getenv("PO_DIR") # directory containing PO pdfs to be processed
    PO_DIR = os.path.normpath(PO_DIR) # for OS flexibility
    #SAVE_DIR = os.path.normpath("SAVE_DIR") # directory to save processed pdfs and excels
    
    # Set up OpenAI API 
    client_openai = OpenAI(api_key=OPENAI_API_KEY)
    openai_config = {
        "model": MODEL,
        "temperature": 0,
    }
    with open("prompt.txt", "r", encoding="utf-8") as f:
        prompt_text = f.read()

    # Set up google API  
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(GOOGLE_JSON_KEY, scope)
    client_gsheet = gspread.authorize(creds)

    # Set up google sheet connections (Main AT WIP & Custom PO Excel)
    sheet_stitchfix = client_gsheet.open(GSHEET_MAIN).worksheet(GSHEET_MAIN_TAB)
    tab_names = {
        'reg': 'Regular',
        'PT': 'Petite',
        'P': 'Plus',
        'reg_00': 'Reg_00',
        'PT_00': 'Petite_00'
    } # add reg-00, PT-00
    sheet_POs = {k: client_gsheet.open(GSHEET_PO).worksheet(tab) for k, tab in tab_names.items()}
    
    # Get list of all PO pdfs
    files = glob.glob(os.path.join(PO_DIR, '*.pdf'))
    files.sort(key=os.path.getmtime)

    try:
        for file_pth in tqdm(files, desc="Processing files", unit="file"):
            
            with open(file_pth, "rb") as f:
                file = client_openai.files.create(
                    file=f,
                    purpose="user_data"
                )
            
            client_input = [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "input_file",
                            "file_id": file.id,
                        },
                        {
                            "type": "input_text",
                            "text": prompt_text
                        }
                    ]
                }
            ]
            
            """PDF to AT WIP (Stitchfix) Google Sheet"""
            response_json = process_PDF(sheet_stitchfix, client_openai, client_input, openai_config)
            #print(response_json, "D")
            
            """AT WIP (Stitchfix) to Custom PO Google Sheet"""
            destination = create_custom_PO(sheet_POs, response_json, creds, tab_names)
            
            # Move pdf to destination directory
            destination_file = destination + "/" + Path(file_pth).name
            if os.path.exists(destination_file):
                os.remove(destination_file)
            shutil.move(file_pth, destination) 

        input("All processes ran. Don't forget to double-check the style#'s. Press enter to exit.")
    except Exception as e:
        print("\nAn error occurred:")
        print(e)
        input("Please rerun script with the same pdf. Press enter to exit.")   
        