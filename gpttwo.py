from openai import OpenAI
import pandas as pd
import gspread
from gspread_dataframe import set_with_dataframe
from oauth2client.service_account import ServiceAccountCredentials
from collections import OrderedDict
import json
import os
import glob
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
from openpyxl import load_workbook
from dotenv import load_dotenv
import shutil
from pathlib import Path

def write_row_to_template2(sheet, row_data, cell_map):
    for field, cell in cell_map.items():
        value = row_data.get(field, "")
        sheet.update(values=[[value]], range_name=cell)

def write_row_to_template(sheet, row_data, cell_map):
    updates = []
    for field, cell in cell_map.items():
        value = row_data.get(field, "")
        updates.append({
            'range': cell,
            'values': [[value]]
        })

    sheet.batch_update(updates)

def get_consistent_output(client, model, temp, client_input, use=False):
    """Call the model twice and only accept consistent, valid output."""
    for _ in range(3):  # Try up to 3 times
        out1 = client.responses.create(model=model, temperature=temp,input=client_input)

        if not use:
            return out1.output_text

        out2 = client.responses.create(model=model, temperature=temp,input=client_input)
        
        if out1.output_text == out2.output_text:
            return out1.output_text
    raise ValueError("Model output inconsistency detected! Try again.")

if __name__ == '__main__':
    load_dotenv() # loads from .env in current directory

    folder_path = os.getenv("STITCHFIX_PTH")
    folder_path = os.path.abspath(folder_path)
    api_key = os.getenv("API_KEY")
    json_key = os.getenv("JSON_KEY")
    packing_order_pth = os.getenv("PACKINGORDER_PTH")
    destination_path = os.getenv("DESTINATION_PTH")
    
    #folder_path = './stitchfix_POs' #'C:/Users/boaz/DataAutomation/Tests'

    # Get list of all files
    files = glob.glob(os.path.join(folder_path, '*.pdf'))#files = glob.glob(os.path.join("./stitchfix_POs", '*'))
    
    client_openai = OpenAI(api_key=api_key)
    model = "gpt-4.1-mini"
    temperature = 0
    
    
    required_cols = [
        "Blank", "Style#", "Color", "Company PO#", "Location", "Vendor", "Size 0",
        "Size 2", "Size 4", "Size 6", "Size 8", "Size 10", "Size 12", "Size 14", "Size 16", "Size 18",
        "Size 20", "Size 22", "Size 24", "Total", "Ordered QTY", "PO ISSUE (Date)",
        "1ST PP REC", "2nd PP", "Requested Ship Date", "Jean Size", "Description"
    ]

    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(json_key, scope)
    client_gsheet = gspread.authorize(creds)

    # Get the most recent file
    #latest_file = max(files, key=os.path.getmtime)
    #print(latest_file)
    
    for file_pth in files:
        
        with open(file_pth, "rb") as f:
            file = client_openai.files.create(
                file=f,
                purpose="user_data"
            )
        client_input = [
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_file",
                        "file_id": file.id,
                    },
                    {
                        "type": "input_text",
                        "text": f"""read the following invoice text and extract the required fields.

                        Always return in valid JSON format.   
                        Do not include commas for the number values. For example, 10000 should not be written as 10,000. The format should be strictly json.
                        Do not put ```json at the beginning or ``` at the end
                        Do not add any explanation, commentary or extra text.
                        The output should be just a json format without any explanations or texts.

                        Extract these fields:
                        - Style# (string of alphabets and digits)
                        - Color (string)
                        - Order# (strings of alphabets and digits)
                        - Sizes (integer value of the quantity; sizes may have characters after (e.g. 0P, 14W, 16W) but they should be just an integer)
                        - Total (integer for the total quantites)
                        - PO ISSUE (Date) (mm/dd),
                        - 1ST PP REC (blank),
                        - 2nd PP (blank),
                        - 2nd PP (blank),
                        - Requested Ship Date (a date mm/dd/yy)
                        - Jean size (may be listed in the "description line item comments"; if not, skip)
                        - Description line item comments (just the product name)

                        Note these facts (do not state this in the output):
                        - ❌ Order# IS NOT the "Release #". Order# may be followed by different location abbreviations like "AZ1", "GA1", "PL1".
                        - The style# comes after the "buyers color descrition" and before "unit price". Do not confuse it with "Item Number".
                        - Style# should be a combination of letters, numbers, and dashes -.
                        - The color should be extracted from the description line.
                        - PO Issue is the PO date.
                        - DO NOT mistaken quantity as item total. Quantity is the value that comes before EACH.
                        - Size comes after "Buyers Item Size Description:"

                        Here is an example:

                        {{
                        "Style#": "WD1017-SF",
                        "Color": "Medium Wash",
                        "Order#": "A1174098-MASTER-US",
                        "Size 0": 2,
                        "Size 2": 5,
                        "Size 4": 10,
                        "Size 6": 12,
                        "Size 8": 8,
                        "Size 10": 7,
                        "Size 12": 11,
                        "Size 14": 4,
                        "Size 16": 3,
                        "Size 18": 6,
                        "Size 20": 1,
                        "Size 22": 0,
                        "Size 24": 9,
                        "Total": 78, 
                        "PO ISSUE (Date)": "1/27",
                        "Requested Ship Date": "6/25/2025",
                        "Jean Size": 31,
                        "Description": "Pocket Wide Leg Jean"
                        }}

                        Here is another for plus sizes:

                        {{
                        "Style#": "WD1017-SF-P",
                        "Color": "Medium Wash",
                        "Order#": "A1174098-AZ1",
                        "Size 14": 4,
                        "Size 16": 3,
                        "Size 18": 6,
                        "Size 20": 1,
                        "Size 22": 0,
                        "Size 24": 9,
                        "Total": 78, 
                        "PO ISSUE (Date)": "1/27",
                        "Requested Ship Date": "6/25/2025",
                        "Jean Size": "",
                        "Description": "Pocket Wide Leg Jean"
                        }}
                        """,
                    },
                ]
            }
        ]
        
        response = get_consistent_output(client_openai, model, temperature, client_input, False)
        #print(response)
        
        response_json = json.loads(response)
        response_json['Company PO#'] = response_json['Order#'].split("-")[0]
        response_json['Location'] = response_json['Order#'].split("-")[1]

        # Note: Map S, M, L to 0, 2, 4, ...

        
        df = pd.DataFrame([response_json])
        df = df.astype(str)
        df = df.reindex(columns=required_cols)

        sheet_stitchfix = client_gsheet.open("Testing").worksheet("Sheet1")

        # Find next empty row
        next_row = len(sheet_stitchfix.get_all_values()) + 1

        # Write data starting from the next empty row
        set_with_dataframe(sheet_stitchfix, df, row=next_row, include_column_header=False)
        
        """StitchFix to Packing Order"""
        sheet_purchaseorder_reg = client_gsheet.open("Packing Order").worksheet("Sheet1")
        sheet_purchaseorder_PT = client_gsheet.open("Packing Order").worksheet("Sheet2")
        sheet_purchaseorder_P = client_gsheet.open("Packing Order").worksheet("Sheet3")
        
        purchase_order_cols = ["Company PO#", "Style#", "PO ISSUE (Date)", "Color", "Size 0", "Size 2", "Size 4", "Size 6", "Size 8", "Size 10", "Size 12", "Size 14", "Size 16", "Size 18", "Size 20", "Size 22", "Size 24", "Total"]
        #df = df.fillna("")
        #purchase_order_dict = df[purchase_order_cols].to_dict(orient="records")[0]
        purchase_order_json = {key: response_json[key] for key in purchase_order_cols if key in response_json}
        #print(purchase_order_dict)
        #print(purchase_order_json)
        cell_map_reg = {
            "Company PO#": "G6", # customer po
            "Style#": "A11",
            "PO ISSUE (Date)": "M6",
            "Color": "A12",
            "Size 0": "C12",
            "Size 2": "D12",
            "Size 4": "E12",
            "Size 6": "F12",
            "Size 8": "G12",
            "Size 10": "H12",
            "Size 12": "I12",
            "Size 14": "J12",
            "Size 16": "K12",
            "Total": "M25"
        }
        cell_map_PT = {
            "Company PO#": "G6", # customer po
            "Style#": "A11",
            "PO ISSUE (Date)": "M6",
            "Color": "A12",
            "Size 0": "C12",
            "Size 2": "D12",
            "Size 4": "E12",
            "Size 6": "F12",
            "Size 8": "G12",
            "Size 10": "H12",
            "Size 12": "I12",
            "Size 14": "J12",
            "Total": "M25"
        }
        cell_map_P = {
            "Company PO#": "G6", # customer po
            "Style#": "A11",
            "PO ISSUE (Date)": "M6",
            "Color": "A12",
            "Size 14": "C12",
            "Size 16": "D12",
            "Size 18": "E12",
            "Size 20": "F12",
            "Size 22": "G12",
            "Size 24": "H12",
            "Total": "M25"
        }

        size = response_json['Style#'].split("-")[-1]
        cell_map = cell_map_reg
        sheet_purchaseorder = sheet_purchaseorder_reg
        sheet_name = 'Sheet1'
        if size in ['PT', 'P']:
            if size == 'PT':
                cell_map = cell_map_PT
                sheet_purchaseorder = sheet_purchaseorder_PT
                sheet_name = 'Sheet2'
            elif size == 'P':
                cell_map = cell_map_P
                sheet_purchaseorder = sheet_purchaseorder_P
                sheet_name = 'Sheet3'

        write_row_to_template(sheet_purchaseorder, purchase_order_json, cell_map)

        # Define the source and destination paths
        #file_name = Path(file_pth)
        source = file_pth
        destination = destination_path

        # Move the file
        destination_file = destination + "/" + Path(source).name
        if os.path.exists(destination_file):
            os.remove(destination_file)
        shutil.move(source, destination)    
        
        if response_json['Location'] == "MASTER":
            """Download as Excel .xlsx"""
            file_id = sheet_purchaseorder.spreadsheet.id
            sheet_to_keep = sheet_name  # the tab name you want to retain

            # 2. Set output path
            output_path = f"{packing_order_pth}/Packing_Order_{response_json['Company PO#']}-{response_json['Location']}.xlsx" #f"C:/Users/boaz/DataAutomation/PackingOrder/Packing_Order_{response_json['Company PO#']}-{response_json['Location']}.xlsx"

            # 3. Download full spreadsheet as Excel
            drive_service = build('drive', 'v3', credentials=creds)
            request = drive_service.files().export_media(
                fileId=file_id,
                mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            fh = io.FileIO(output_path, 'wb')
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
                print("Download progress: {:.0f}%".format(status.progress() * 100))
            fh.close()

            blank_json = {key: '' for key in purchase_order_cols if key in response_json}

            write_row_to_template(sheet_purchaseorder, blank_json, cell_map)

            # 4. Remove extra sheets
            wb = load_workbook(output_path)
            for s in wb.sheetnames:
                if s != sheet_to_keep:
                    wb.remove(wb[s])
            wb.save(output_path)

    
        