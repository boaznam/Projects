from openai import OpenAI
import pandas as pd
import gspread
from gspread_dataframe import set_with_dataframe
from oauth2client.service_account import ServiceAccountCredentials
from collections import OrderedDict
import json
import os
import glob
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
from openpyxl import load_workbook
from dotenv import load_dotenv
import shutil
from pathlib import Path

def write_to_gsheet(sheet, data, cell_map):
    """
    Writes values from row_data to specific cells in the sheet using cell_map.
    Used to write 'AT WIP' gsheet values into custom PO gsheet.

    Args:
        sheet (gsheet object): gspread worksheet to write into.
        data (dict): mapping of field names (e.g. "CompanyPO#") and their values.
        cell_map (dict): mapping of field names to cell addresses.
    
    Returns:
        None
    """
    updates = []
    for field, cell in cell_map.items():
        value = data.get(field, "")
        updates.append({
            'range': cell,
            'values': [[value]]
        })

    sheet.batch_update(updates)

def write_to_excel(file_pth, PO_json, cell_map):
    """
    Write data to an existing Excel file using cell mappings.
    Used to write location PO info from 'AT WIP' to previously saved custom PO excel.

    Args:
        file_path (str): Path to the Excel file.
        purchase_order_json (dict): Data to write (e.g., {"Company PO#": "12345", "0": 10}).
        cell_map (dict): mapping of field names to cell addresses (e.g., {"Company PO#": "G6", "0": "C12"}).
    """
    try:
        # Load the existing workbookt
        try:
            workbook = load_workbook(file_pth)
        except FileNotFoundError:
            raise ValueError(f"File {file_pth} not found. The file does not exist in \
                             given path. It may have been deleted or the MASTER pdf \
                             may have not been processed yet.")
        
        # Access the worksheet (there should be only one sheet (tab) named after the size)
        tab = workbook.active
        
        # Write data to the specified cells
        for key, value in PO_json.items():
            if key in cell_map:
                cell_address = cell_map[key]
                tab[cell_address] = value
        
        # Save the workbook
        workbook.save(file_pth)
        print(f"Data successfully written to {file_pth}.")
        
    except Exception as e:
        print(f"Error writing to Excel: {e}")

def get_consistent_output(client, client_input, model, temperature, use=False):
    """Call the model twice and only accept consistent, valid output."""
    for _ in range(3):  # Try up to 3 times
        out1 = client.responses.create(model=model, temperature=temperature,input=client_input)

        if not use:
            return out1.output_text

        out2 = client.responses.create(model=model, temperature=temperature,input=client_input)
        
        if out1.output_text == out2.output_text:
            return out1.output_text
    raise ValueError("Model output inconsistency detected! Try again.")

def process_PDF(sheet, client_openai, client_input, openai_config):
    """
    Runs openai's model to process a pdf and the prompt to extract
    necessary data as a json and write into the given gsheet.
    Used to write pdf data into 'AT WIP'.

    Note:
        Extract Company PO# and Location from Order#
        Change Requested Ship Date to CXL
        Change Jean Size to Inseam
        Change 'Size {i}' to '{i}'; 16 -> 16/1x, 18 -> 18/2x, 20 -> 20/3x

    Args:
        sheet (gsheet object): google sheet to input pdf data to.
        client_open (openai object): OpenAI(api_key=OPENAI_API_KEY).
        client_input (dict): input to the openai model.
        openai_config (dict): configs like model and temperature.
    
    Returns:
        openai model's output as json "response_json"
    """
    response = get_consistent_output(client_openai, client_input, **openai_config) # Set use=True for inconsistent output
    
    response_json = json.loads(response)
    response_json['Company PO#'] = response_json['Order#'].split("-")[0]
    #response_json['Location'] = response_json['Order#'].split("-")[1]
    response_json['CXL'] = response_json.pop('Requested Ship Date')
    response_json['Inseam'] = response_json.pop('Jean Size')

    for size in range(0, 26, 2):
        if size == 16:
            response_json[f'{size}/1x'] = response_json.pop(f'Size {size}', None)
        elif size == 18:
            response_json[f'{size}/2x'] = response_json.pop(f'Size {size}', None)
        elif size == 20:
            response_json[f'{size}/3x'] = response_json.pop(f'Size {size}', None)
        else:
            response_json[f'{size}'] = response_json.pop(f'Size {size}', None)
    
    response_json = {k: ("" if v is None else v) for k, v in response_json.items()} # replace None with ""
    
    required_cols = sheet.row_values(1)
    required_cols = [word.strip() for word in required_cols]
    """
    Example columns from main sheet:
    required_cols = [
        "Style#", "Color", "Company PO#", "Location", "Vendor", "0",
        "2", "4", "6", "8", "10", "12", "14", "16", "18",
        "20", "22", "24", "Total", "Ordered QTY", "PO ISSUE",
        "1ST PP REC", "2nd PP", "Requested Ship Date", "Jean Size", "Description", "Fabric", "C Price"
    ]
    """
    df = pd.DataFrame([response_json])
    df = df.astype(str)
    df = df.reindex(columns=required_cols) # extracts all columns from main google sheet

    next_row = len(sheet.get_all_values()) + 1 # find next empty row

    set_with_dataframe(sheet, df, row=next_row, include_column_header=False)

    return response_json

def create_custom_PO(sheet, response_json, creds, tab_names):
    """
    Maps 'AT WIP' gsheet values to custom PO gsheet.
    Custom PO excel is saved only for MASTER PDF. 
        Location PDFs GA1, PL1, AZ1 are written into it at later dates. We assume MASTER always comes first.

    Note:
        - PDFs of different sizes have slightly different formats. Map accordingly.
        - MASTER and location PDFs are written into different rows. 

    Args:
        sheet (gsheet object): google sheet to input 'AT WIP' data to.
        response_json (dict): output of openai model with all necessary values.
        creds: google drive API credentials to access the API for exports gsheets as excel.

    Returns:
        destination (directory path where custom PO excel is saved)
    """
    size = response_json["STYLE#"].split("-")[-1]
    location = response_json["Location"]

    # Directory path to store pdfs and excel
    destination = f"../{response_json["STYLE#"]}/PO/{response_json["Company PO#"]}_{response_json["STYLE#"]}"
    
    # Check that Master file is created before any other location PDFs
    if not os.path.exists(destination):
        if location != "MASTER":
            raise ValueError(f"Master file has not been processed yet. Current PO is for location: {location}")
        os.makedirs(destination, exist_ok=False)
 
    if size not in ['PT', 'P']:
        size = 'reg'

    # Fields that are required from MASTER PDFs of all sizes.
    common_fields = {
        "Company PO#": "G6",
        "DESCRIPTION": "A10",
        "STYLE#": "A11",
        "PO ISSUE": "M6",
        "CXL": "M8",
        "Color": "A12",
        "Total": "M25",
        "C Price": "N12"
    }

    # Create mapping to columns for each size.
    size_col_map = {
        'reg': {"0": "C", "2": "D", "4": "E", "6": "F", "8": "G", "10": "H", "12": "I", "14": "J", "16": "K"},
        'PT': {"0": "C", "2": "D", "4": "E", "6": "F", "8": "G", "10": "H", "12": "I", "14": "J"},
        'P': {"14": "C", "16": "D", "18": "E", "20": "F", "22": "G", "24": "H"}
    }

    # Create mapping to rows based on location
    loc_row_map = {
        'MASTER': 12,
        'GA1': 14,
        'AZ1': 16,
        'PL1': 18 
    }
    row = loc_row_map[location]

    # Create complete cell mappings
    """ 
    Example for GA1:
    cell_maps = {
        'reg': {
            "0": "C14", "2": "D14", "4": "E14", "6": "F14",
            "8": "G14", "10": "H14", "12": "I14", "14": "J14", "16": "K14"
        },
        'PT': {
            "0": "C14", "2": "D14", "4": "E14", "6": "F14",
            "8": "G14", "10": "H14", "14": "I14", "14": "J14"
        },
        'P': {
            "14": "C14", "16": "D14", "18": "E14", "20": "F14",
            "22": "G14", "24": "H14"
        }
    }
    """
    cell_maps = {
        size: {
            **(common_fields if location == "MASTER" else {}),
            **{key: f"{col}{row}" for key, col in col_map.items()}
        } for size, col_map in size_col_map.items()
    }
    
    PO_cols = cell_maps[size].keys()
    PO_json = {key: response_json[key] for key in PO_cols if key in response_json}

    # Write into gsheets
    sheet_PO = sheet[size]
    cell_map = cell_maps[size]
    sheet_name = tab_names[size]
    
    # Only Master pdf is saved as excel. Location pdfs write into the saved excel.
    if response_json['Location'] == "MASTER":

        write_to_gsheet(sheet_PO, PO_json, cell_map)

        """Download as Excel .xlsx"""
        file_id = sheet_PO.spreadsheet.id
        sheet_to_keep = sheet_name  # the tab name you want to retain

        # Set output path
        output_path = f"{destination}/PO_{response_json['Company PO#']}-{response_json['Location']}.xlsx" 

        # Download full spreadsheet as Excel
        drive_service = build('drive', 'v3', credentials=creds)
        request = drive_service.files().export_media(
            fileId=file_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        fh = io.FileIO(output_path, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
            print("Download progress: {:.0f}%".format(status.progress() * 100))
        fh.close()

        # Clear gsheet cells
        blank_json = {key: '' for key in PO_cols if key in response_json}
        write_to_gsheet(sheet_PO, blank_json, cell_map)

        # Remove extra sheets (tabs)
        wb = load_workbook(output_path)
        for s in wb.sheetnames:
            if s != sheet_to_keep:
                wb.remove(wb[s])
        wb.save(output_path)

    else:
        # check if PO excel exists from Master
        file_pth=glob.glob(os.path.join(destination, '*.xlsx'))[0]#path to .xlsx
        write_to_excel(file_pth, PO_json, cell_map)

    return destination



if __name__ == '__main__':

    # Load from .env in current directory
    load_dotenv()
    
    # Load keys and directory paths
    OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
    MODEL = os.getenv("MODEL")
    #TEMPERATURE = os.getenv("TEMPERATURE")
    GOOGLE_JSON_KEY = os.getenv("GOOGLE_JSON_KEY")
    GSHEET_MAIN, GSHEET_MAIN_TAB = os.getenv("GSHEET_MAIN"), os.getenv("GSHEET_MAIN_TAB")
    GSHEET_PO = os.getenv("GSHEET_PO")
    ROOT_DIR = os.getenv("ROOT_DIR")
    PO_DIR = os.getenv("PO_DIR") # directory containing PO pdfs to be processed
    PO_DIR = os.path.normpath(PO_DIR) # for OS flexibility
    #SAVE_DIR = os.path.normpath("SAVE_DIR") # directory to save processed pdfs and excels
    
    # Set up OpenAI API 
    client_openai = OpenAI(api_key=OPENAI_API_KEY)
    openai_config = {
        "model": MODEL,
        "temperature": 0,
    }
    with open("prompt.txt", "r", encoding="utf-8") as f:
        prompt_text = f.read()

    # Set up google API  
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(GOOGLE_JSON_KEY, scope)
    client_gsheet = gspread.authorize(creds)

    # Set up google sheet connections (Main AT WIP & Custom PO Excel)
    sheet_stitchfix = client_gsheet.open(GSHEET_MAIN).worksheet(GSHEET_MAIN_TAB)
    tab_names = {
        'reg': 'Regular',
        'PT': 'Petite',
        'P': 'Plus'
    }
    sheet_POs = {k: client_gsheet.open(GSHEET_PO).worksheet(tab) for k, tab in tab_names.items()}
    
    # Get list of all PO pdfs
    files = glob.glob(os.path.join(PO_DIR, '*.pdf'))
    
    for file_pth in files:
        
        with open(file_pth, "rb") as f:
            file = client_openai.files.create(
                file=f,
                purpose="user_data"
            )
        
        client_input = [
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_file",
                        "file_id": file.id,
                    },
                    {
                        "type": "input_text",
                        "text": prompt_text
                    }
                ]
            }
        ]
        
        """PDF to AT WIP (Stitchfix) Google Sheet"""
        response_json = process_PDF(sheet_stitchfix, client_openai, client_input, openai_config)
        print(response_json, "D")

        """AT WIP (Stitchfix) to Custom PO Google Sheet"""
        """
        destination = create_custom_PO(sheet_POs, response_json, creds, tab_names)
        
        # Move pdf to destination directory
        destination_file = destination + "/" + Path(file_pth).name
        if os.path.exists(destination_file):
            os.remove(destination_file)
        shutil.move(file_pth, destination)    
        """